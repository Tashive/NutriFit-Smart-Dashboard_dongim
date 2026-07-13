"""
엑셀 파일 시트 검사 스크립트

이 스크립트는 `성분_수작업입력_대상.xlsx`와 `성분_수작업입력_초안.xlsx` 파일의
상위 10개 행 셀 값을 각각 출력하여 헤더 위치 및 데이터 상태를 검사합니다.
"""

import os
import sys
from openpyxl import load_workbook

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
TARGET_PATH = os.path.join(DATA_DIR, "성분_수작업입력_대상.xlsx")
DRAFT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_초안.xlsx")

def dump_sheet(path):
    print(f"\n===== File: {os.path.basename(path)} =====")
    if not os.path.exists(path):
        print("파일이 존재하지 않습니다.")
        return
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    for r in range(1, 12):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        print(f"Row {r:2d}: {row_vals}")

dump_sheet(TARGET_PATH)
dump_sheet(DRAFT_PATH)
