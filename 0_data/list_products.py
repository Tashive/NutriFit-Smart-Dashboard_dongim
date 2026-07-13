"""
제품명 목록 및 인덱스 확인 스크립트

이 스크립트는 `성분_수작업입력_대상.xlsx` 파일을 로드하여
각 행 번호, 플랫폼, 제품명을 출력합니다. 이를 통해 각 제품의
카테고리와 성분을 분류하는 로직을 구상합니다.
"""

import os
import sys
from openpyxl import load_workbook

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
EXCEL_PATH = os.path.join(DATA_DIR, "성분_수작업입력_대상.xlsx")

wb = load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

# 4행부터 데이터 시작 (3행은 헤더)
for row in range(4, 49):
    platform = ws.cell(row=row, column=1).value
    name = ws.cell(row=row, column=2).value
    print(f"Row {row} | {platform} | {name}")
