"""
추가분 콜라겐 및 유산균 상품 추출 및 엑셀 생성 스크립트 (비율 조정 버전)

이 스크립트는 `NutriFit-Dashboard-Data-file.csv` 파일에서 콜라겐 및 유산균 관련
신규 상품을 각각 최대 15개씩 추출합니다.
- 기존 스크립트 문제 해결: 데이터가 누적된 순서에 따라 Coupang 상품만 15개 채워지는 편향을 제거합니다.
- 플랫폼별 매칭 비중을 반영하여 비례 배분합니다:
  * 유산균 (총 15개): iHerb 11개, OliveYoung 3개, Coupang 1개
  * 콜라겐 (총 15개): iHerb 11개, OliveYoung 3개, Coupang 1개
- 검색 대상 제외: 기존 `성분_수작업입력_초안.xlsx`에 포함된 상품(제품명 기준)
- 최종 결과물은 `성분_수작업입력_추가분_콜라겐유산균.xlsx` 파일로 덮어씌워 새로 저장됩니다.
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
CSV_PATH = os.path.join(DATA_DIR, "NutriFit-Dashboard-Data-file.csv")
EXCEL_DRAFT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_초안.xlsx")
OUTPUT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_추가분_콜라겐유산균.xlsx")

def clean_price_val(val):
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    if '.' in val_str:
        val_str = val_str.split('.')[0]
    cleaned = ''.join([c for c in val_str if c.isdigit()])
    return int(cleaned) if cleaned else None

def main():
    if not os.path.exists(CSV_PATH) or not os.path.exists(EXCEL_DRAFT_PATH):
        print("필요한 데이터 파일이 존재하지 않습니다.")
        return

    # 1. 기존 제품명 로드 (제외 대상)
    df_draft = pd.read_excel(EXCEL_DRAFT_PATH, skiprows=3)
    existing_names = set(df_draft["제품명"].dropna().str.strip())

    # 2. 전체 CSV 데이터 로드
    df_csv = pd.read_csv(CSV_PATH, low_memory=False)

    # 3. 키워드 정의
    collagen_kws = ["콜라겐", "collagen", "이너뷰티"]
    probiotic_kws = ["유산균", "프로바이오틱", "probiotic", "프로바이오틱스", "락토", "lacto", "비피더스", "bifidobacterium"]

    # 각 플랫폼별 매칭 상품을 저장할 딕셔너리
    collagen_by_platform = {"Coupang": [], "OliveYoung": [], "iHerb": []}
    probiotic_by_platform = {"Coupang": [], "OliveYoung": [], "iHerb": []}

    name_cols = {"Coupang": "product_name", "OliveYoung": "name", "iHerb": "displayName"}
    price_cols = {"Coupang": "price", "OliveYoung": "price_cur", "iHerb": "discountPrice"}

    seen_names = set()

    for _, row in df_csv.iterrows():
        platform = row["platform"]
        if pd.isna(platform) or platform not in name_cols:
            continue
        name_col = name_cols[platform]
        price_col = price_cols[platform]

        name = row.get(name_col)
        if pd.isna(name):
            continue
        name_str = str(name).strip()

        # 이미 존재하는 제품 또는 중복 수집 대상 제외
        if name_str in existing_names or name_str in seen_names:
            continue

        name_lower = name_str.lower()
        price = clean_price_val(row.get(price_col))

        product_info = {
            "플랫폼": platform,
            "제품명": name_str,
            "가격": price
        }

        is_collagen = any(kw in name_lower for kw in collagen_kws)
        is_probiotic = any(kw in name_lower for kw in probiotic_kws)

        if is_collagen:
            collagen_by_platform[platform].append(product_info)
            seen_names.add(name_str)
        elif is_probiotic:
            probiotic_by_platform[platform].append(product_info)
            seen_names.add(name_str)

    # 비례 배분 기준에 맞춰 상품 추출
    # 유산균 (총 15개): iHerb 11개, OliveYoung 3개, Coupang 1개
    probiotic_final = (
        probiotic_by_platform["iHerb"][:11] +
        probiotic_by_platform["OliveYoung"][:3] +
        probiotic_by_platform["Coupang"][:1]
    )

    # 콜라겐 (총 15개): iHerb 11개, OliveYoung 3개, Coupang 1개
    collagen_final = (
        collagen_by_platform["iHerb"][:11] +
        collagen_by_platform["OliveYoung"][:3] +
        collagen_by_platform["Coupang"][:1]
    )

    print(f"추출된 신규 콜라겐 상품 수: {len(collagen_final)}건 (Coupang: {len(collagen_by_platform['Coupang'][:1])}, OliveYoung: {len(collagen_by_platform['OliveYoung'][:3])}, iHerb: {len(collagen_by_platform['iHerb'][:11])})")
    print(f"추출된 신규 유산균 상품 수: {len(probiotic_final)}건 (Coupang: {len(probiotic_by_platform['Coupang'][:1])}, OliveYoung: {len(probiotic_by_platform['OliveYoung'][:3])}, iHerb: {len(probiotic_by_platform['iHerb'][:11])})")

    # 4. 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "추가분 입력"
    ws.views.sheetView[0].showGridLines = True

    # 스타일 설정
    font_name = "맑은 고딕"
    font_normal = Font(name=font_name, size=10)
    font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    font_banner = Font(name=font_name, size=11, bold=True, color="7F3F00")

    fill_banner = PatternFill(start_color="FFEAD2", end_color="FFEAD2", fill_type="solid")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    border_thin = Side(border_style="thin", color="D9D9D9")
    border_medium = Side(border_style="medium", color="1F4E78")

    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # 1행: 안내 배너
    ws.merge_cells("A1:E1")
    ws["A1"] = "★ 안내 문구: 카테고리는 [멀티비타민 / 비타민C / 오메가3 / 유산균 / 콜라겐 / 마그네슘] 중 하나로 채워주세요."
    ws["A1"].font = font_banner
    ws["A1"].fill = fill_banner
    ws["A1"].alignment = align_center
    ws["A1"].border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    ws.row_dimensions[1].height = 35

    # 2행~3행: 공백 패딩
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15

    # 4행: 헤더
    headers = ["플랫폼", "제품명", "가격", "카테고리", "성분"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = Border(left=border_thin, right=border_thin, top=border_medium, bottom=border_medium)
    ws.row_dimensions[4].height = 28

    # 데이터 쓰기 (콜라겐 15개 먼저, 유산균 15개 그 다음)
    all_data = []
    for item in collagen_final:
        all_data.append((item, "콜라겐"))
    for item in probiotic_final:
        all_data.append((item, "유산균"))

    for i, (item, cat_memo) in enumerate(all_data):
        r_idx = 5 + i
        use_zebra = (r_idx % 2 == 0)

        # 플랫폼 (A)
        c_a = ws.cell(row=r_idx, column=1, value=item["플랫폼"])
        c_a.alignment = align_center

        # 제품명 (B)
        c_b = ws.cell(row=r_idx, column=2, value=item["제품명"])
        c_b.alignment = align_left

        # 가격 (C)
        c_c = ws.cell(row=r_idx, column=3, value=item["가격"])
        c_c.alignment = align_right
        c_c.number_format = "₩#,##0"

        # 카테고리 (D) - 빈칸이나 예상 카테고리 메모 추가
        c_d = ws.cell(row=r_idx, column=4, value="")
        c_d.alignment = align_center
        c_d.comment = Comment(cat_memo, "System")

        # 성분 (E) - 빈칸
        c_e = ws.cell(row=r_idx, column=5, value="")
        c_e.alignment = align_center

        # 테두리 및 스타일 설정
        for col_num in range(1, 6):
            c = ws.cell(row=r_idx, column=col_num)
            c.font = font_normal
            c.border = cell_border
            if use_zebra:
                c.fill = fill_zebra

        ws.row_dimensions[r_idx].height = 22

    # 열 너비 설정
    widths = {
        "A": 15,
        "B": 65,
        "C": 15,
        "D": 22,
        "E": 22
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT_PATH)
    print("--- 추가분 엑셀 생성 완료 ---")
    print(f"저장 경로: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
