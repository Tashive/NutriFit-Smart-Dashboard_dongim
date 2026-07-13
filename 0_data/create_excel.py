"""
NutriFit 수작업 성분 입력용 엑셀 생성 스크립트

이 스크립트는 `NutriFit-Dashboard-Data-file.csv` 파일을 읽어
각 플랫폼(Coupang, OliveYoung, iHerb)별로 상위 15개 제품을 추출합니다.
추출된 총 45개 제품에 대해 지정된 컬럼(플랫폼, 제품명, 가격, 카테고리, 성분)으로 구성하고
안내 문구와 디자인 스타일을 적용하여 `성분_수작업입력_대상.xlsx` 엑셀 파일로 저장합니다.
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
CSV_PATH = os.path.join(DATA_DIR, "NutriFit-Dashboard-Data-file.csv")
OUTPUT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_대상.xlsx")

def clean_price(val):
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    # 숫자 이외의 문자 제거 (₩, 콤마 등)
    cleaned = "".join([c for c in val_str if c.isdigit()])
    if cleaned:
        # 소수점이 포함된 경우 정수형 변환을 위해 float 거치기
        return int(float(cleaned))
    return None

def main():
    if not os.path.exists(CSV_PATH):
        print(f"Error: {CSV_PATH}가 존재하지 않습니다.")
        return

    df = pd.read_csv(CSV_PATH, low_memory=False)
    
    # 각 플랫폼별 상위 15개 상품 추출
    platforms = ["Coupang", "OliveYoung", "iHerb"]
    extracted_data = []

    for platform in platforms:
        group = df[df["platform"] == platform]
        top15 = group.head(15).copy()
        
        # 플랫폼별 컬럼 매핑 설정
        if platform == "Coupang":
            name_col = "product_name"
            price_col = "price"
        elif platform == "OliveYoung":
            name_col = "name"
            price_col = "price_cur"
        else: # iHerb
            name_col = "displayName"
            price_col = "discountPrice"
            
        for _, row in top15.iterrows():
            name = row.get(name_col)
            price_raw = row.get(price_col)
            price = clean_price(price_raw)
            
            extracted_data.append({
                "플랫폼": platform,
                "제품명": name,
                "가격": price,
                "카테고리": "",
                "성분": ""
            })
            
    df_out = pd.DataFrame(extracted_data)
    
    # openpyxl을 이용한 엑셀 파일 생성 및 포맷팅
    wb = Workbook()
    ws = wb.active
    ws.title = "성분 수작업 입력"
    
    # 그리드라인 보이기 설정
    ws.views.sheetView[0].showGridLines = True
    
    # 폰트 및 스타일 정의
    font_name = "맑은 고딕"
    font_normal = Font(name=font_name, size=10)
    font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    font_banner = Font(name=font_name, size=11, bold=True, color="7F3F00") # 진한 주황색 글씨
    
    fill_banner = PatternFill(start_color="FFEAD2", end_color="FFEAD2", fill_type="solid") # 연한 주황 배경
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # 스틸 블루
    fill_zebra = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid") # 짝수행용 밝은 스틸 블루/회색
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_medium = Side(border_style="medium", color="1F4E78")
    
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # 1행 안내 문구 작성 및 병합
    ws.merge_cells("A1:E1")
    ws["A1"] = "★ 안내 문구: 카테고리는 [멀티비타민 / 비타민C / 오메가3 / 유산균 / 콜라겐 / 마그네슘] 중 하나로 채워주세요."
    ws["A1"].font = font_banner
    ws["A1"].fill = fill_banner
    ws["A1"].alignment = align_center
    ws["A1"].border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    ws.row_dimensions[1].height = 35
    
    # 2행은 공백
    ws.row_dimensions[2].height = 15
    
    # 3행 헤더 작성
    headers = ["플랫폼", "제품명", "가격", "카테고리", "성분"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = Border(left=border_thin, right=border_thin, top=border_medium, bottom=border_medium)
    ws.row_dimensions[3].height = 28
    
    # 4행부터 데이터 채우기
    for r_idx, row_data in enumerate(df_out.itertuples(index=False), 4):
        # 짝수행 제브라 스트라이프
        use_zebra = (r_idx % 2 == 0)
        
        # 플랫폼 (A)
        cell_a = ws.cell(row=r_idx, column=1, value=row_data.플랫폼)
        cell_a.alignment = align_center
        
        # 제품명 (B)
        cell_b = ws.cell(row=r_idx, column=2, value=row_data.제품명)
        cell_b.alignment = align_left
        
        # 가격 (C)
        cell_c = ws.cell(row=r_idx, column=3, value=row_data.가격)
        cell_c.alignment = align_right
        cell_c.number_format = "₩#,##0"
        
        # 카테고리 (D)
        cell_d = ws.cell(row=r_idx, column=4, value="")
        cell_d.alignment = align_center
        
        # 성분 (E)
        cell_e = ws.cell(row=r_idx, column=5, value="")
        cell_e.alignment = align_center
        
        # 모든 데이터 셀에 폰트 및 테두리 적용
        for col_num in range(1, 6):
            c = ws.cell(row=r_idx, column=col_num)
            c.font = font_normal
            c.border = cell_border
            if use_zebra:
                c.fill = fill_zebra
                
        ws.row_dimensions[r_idx].height = 22
        
    # 열 너비 설정
    widths = {
        "A": 15,
        "B": 65,
        "C": 15,
        "D": 22,
        "E": 22
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save(OUTPUT_PATH)
    print("--- 엑셀 파일 생성 완료 ---")
    print(f"저장 경로: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
