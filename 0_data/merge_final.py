"""
최종 성분 통합 엑셀 생성 및 카테고리 집계 스크립트

이 스크립트는 두 개의 초안 엑셀 파일을 하나로 통합하여 `성분_통합_최종.xlsx` 파일을 생성합니다:
1. `성분_수작업입력_초안.xlsx` (카테고리가 '확인필요' 또는 공란이 아닌 유효 행만 필터링, 컬럼명 '가격(원)' 매핑)
2. `성분_수작업입력_추가분_초안.xlsx` (전체 30개 행 모두 포함)
통합 결과물은 기존 엑셀과 동일한 일관성 있는 디자인 테마(스틸 블루 헤더, 연주황 안내 배너, 제브라 스트라이프 등)가
적용되며, 저장 후 카테고리별 상품 개수를 집계하여 터미널에 출력합니다.
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
DRAFT1_PATH = os.path.join(DATA_DIR, "성분_수작업입력_초안.xlsx")
DRAFT2_PATH = os.path.join(DATA_DIR, "성분_수작업입력_추가분_초안.xlsx")
OUTPUT_PATH = os.path.join(DATA_DIR, "성분_통합_최종.xlsx")

def main():
    if not os.path.exists(DRAFT1_PATH) or not os.path.exists(DRAFT2_PATH):
        print("필요한 초안 파일이 누락되었습니다.")
        return

    # 1. 첫 번째 초안 로드 및 필터링 (skiprows=3로 헤더가 4행)
    df1 = pd.read_excel(DRAFT1_PATH, skiprows=3)
    df1_filtered = df1[df1["카테고리"].notna() & (df1["카테고리"] != "확인필요")].copy()
    # 가격(원) 컬럼명을 가격으로 통일
    df1_filtered = df1_filtered.rename(columns={"가격(원)": "가격"})
    
    # 2. 두 번째 초안 로드 (전체 30행)
    df2 = pd.read_excel(DRAFT2_PATH, skiprows=3)
    
    # 두 데이터 프레임의 필요한 컬럼만 추출하여 병합
    cols = ["플랫폼", "제품명", "가격", "카테고리", "성분"]
    df1_subset = df1_filtered[cols]
    df2_subset = df2[cols]
    
    df_combined = pd.concat([df1_subset, df2_subset], ignore_index=True)
    print(f"초안 1 유효 데이터: {len(df1_subset)}건")
    print(f"초안 2 추가 데이터: {len(df2_subset)}건")
    print(f"통합 최종 데이터: {len(df_combined)}건")

    # 3. openpyxl로 최종 엑셀 쓰기
    wb = Workbook()
    ws = wb.active
    ws.title = "최종 통합 성분"
    ws.views.sheetView[0].showGridLines = True
    
    # 스타일 설정
    font_name = "맑은 고딕"
    font_normal = Font(name=font_name, size=10)
    font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    font_banner = Font(name=font_name, size=11, bold=True, color="7F3F00")

    fill_banner = PatternFill(start_color="FFEAD2", end_color="FFEAD2", fill_type="solid")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    border_thin = Side(border_style="thin", color="D9D9D9")
    border_medium = Side(border_style="medium", color="1F4E78")

    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # 1행: 안내 배너
    ws.merge_cells("A1:E1")
    ws["A1"] = "★ 안내 문구: 카테고리는 [멀티비타민 / 비타민C / 오메가3 / 유산균 / 콜라겐 / 마그네슘] 중 하나로 채워주세요."
    ws["A1"].font = font_banner
    ws["A1"].fill = fill_banner
    ws["A1"].alignment = align_center
    ws["A1"].border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    ws.row_dimensions[1].height = 35

    # 2행~3행: 공백 패딩
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15

    # 4행: 헤더
    headers = ["플랫폼", "제품명", "가격", "카테고리", "성분"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = Border(left=border_thin, right=border_thin, top=border_medium, bottom=border_medium)
    ws.row_dimensions[4].height = 28

    # 5행부터 데이터 기록
    for i, row_data in enumerate(df_combined.itertuples(index=False)):
        r_idx = 5 + i
        use_zebra = (r_idx % 2 == 0)

        # 플랫폼 (A)
        c_a = ws.cell(row=r_idx, column=1, value=row_data.플랫폼)
        c_a.alignment = align_center

        # 제품명 (B)
        c_b = ws.cell(row=r_idx, column=2, value=row_data.제품명)
        c_b.alignment = align_left

        # 가격 (C)
        # float 형태인 경우 정수로 변환하여 깔끔하게 표시
        price_val = row_data.가격
        if pd.notna(price_val):
            price_val = int(price_val)
        c_c = ws.cell(row=r_idx, column=3, value=price_val)
        c_c.alignment = align_right
        c_c.number_format = "₩#,##0"

        # 카테고리 (D)
        c_d = ws.cell(row=r_idx, column=4, value=row_data.카테고리)
        c_d.alignment = align_center

        # 성분 (E)
        c_e = ws.cell(row=r_idx, column=5, value=row_data.성분)
        c_e.alignment = align_center

        # 스타일 및 보더 설정
        for col_num in range(1, 6):
            c = ws.cell(row=r_idx, column=col_num)
            c.font = font_normal
            c.border = cell_border
            if use_zebra:
                c.fill = fill_zebra

        ws.row_dimensions[r_idx].height = 22

    # 열 너비 설정
    widths = {
        "A": 15,
        "B": 65,
        "C": 15,
        "D": 22,
        "E": 22
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT_PATH)
    print("--- 통합 최종 엑셀 생성 완료 ---")
    print(f"저장 경로: {OUTPUT_PATH}")

    # 카테고리별 집계 결과 출력
    print("\n=== 최종 통합본 카테고리별 상품 수 ===")
    counts = df_combined["카테고리"].value_counts()
    print(counts.to_string())

if __name__ == "__main__":
    main()
