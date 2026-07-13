"""
대상 엑셀 파일 행 수 및 가격 정보 출력 스크립트

이 스크립트는 `성분_수작업입력_대상.xlsx` 파일의 모든 행을 읽어
행 번호, 플랫폼, 제품명, 가격을 출력하고 총 행 수를 확인합니다.
"""

import os
import sys
from openpyxl import load_workbook

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
TARGET_PATH = os.path.join(DATA_DIR, "성분_수작업입력_대상.xlsx")

wb = load_workbook(TARGET_PATH, data_only=True)
ws = wb.active

print(f"Max Row: {ws.max_row}")
for r in range(1, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
    print(f"Row {r:2d}: {row_vals}")
