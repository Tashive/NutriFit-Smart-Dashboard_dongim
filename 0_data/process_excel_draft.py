"""
수작업 성분 입력용 엑셀 초안 생성 스크립트

이 스크립트는 `성분_수작업입력_대상.xlsx` 파일을 복사/로드하여,
각 제품명의 텍스트 정보를 바탕으로 1차 카테고리(멀티비타민, 비타민C, 오메가3, 유산균, 콜라겐, 마그네슘)
및 구체적 성분을 채워 넣습니다. 규칙에 따라 확실하지 않거나 6대 카테고리에 속하지 않는 제품은
"확인필요"로 기록하며, 최종 결과물을 `성분_수작업입력_초안.xlsx`로 저장합니다.
"""

import os
import sys
from openpyxl import load_workbook

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
INPUT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_대상.xlsx")
OUTPUT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_초안.xlsx")

# 행 번호별 분류 매핑 정의 (1-indexed 행 번호 기준)
MAPPING = {
    4:  {"category": "마그네슘", "ingredient": "마그네슘"},
    5:  {"category": "확인필요", "ingredient": "엽산, 비타민 B-12"},
    6:  {"category": "확인필요", "ingredient": "케르세틴, 브로멜라인"},
    7:  {"category": "확인필요", "ingredient": "확인필요"},
    8:  {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    9:  {"category": "확인필요", "ingredient": "비타민 B"},
    10: {"category": "확인필요", "ingredient": "철분"},
    11: {"category": "확인필요", "ingredient": "확인필요"},
    12: {"category": "마그네슘", "ingredient": "마그네슘 시트레이트"},
    13: {"category": "유산균",   "ingredient": "프로바이오틱스, 유산균"},
    14: {"category": "확인필요", "ingredient": "베르베린"},
    15: {"category": "확인필요", "ingredient": "비타민 B-12"},
    16: {"category": "확인필요", "ingredient": "확인필요"},
    17: {"category": "멀티비타민", "ingredient": "비타민"},
    18: {"category": "확인필요", "ingredient": "L-리신"},
    19: {"category": "확인필요", "ingredient": "확인필요"},
    20: {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    21: {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    22: {"category": "확인필요", "ingredient": "포타슘, 칼륨"},
    23: {"category": "멀티비타민", "ingredient": "비타민"},
    24: {"category": "비타민C",   "ingredient": "비타민C"},
    25: {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    26: {"category": "확인필요", "ingredient": "확인필요"},
    27: {"category": "확인필요", "ingredient": "확인필요"},
    28: {"category": "멀티비타민", "ingredient": "비타민, 아르기닌"},
    29: {"category": "확인필요", "ingredient": "포타슘, 칼륨"},
    30: {"category": "멀티비타민", "ingredient": "비타민"},
    31: {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    32: {"category": "비타민C",   "ingredient": "비타민C"},
    33: {"category": "확인필요", "ingredient": "비타민C, 비타민D"},
    34: {"category": "유산균",   "ingredient": "프로바이오틱"},
    35: {"category": "확인필요", "ingredient": "비타민D3, 비타민K2"},
    36: {"category": "마그네슘", "ingredient": "마그네슘 리시네이트 글리시네이트"},
    37: {"category": "마그네슘", "ingredient": "마그네슘 글리시네이트"},
    38: {"category": "확인필요", "ingredient": "비타민D3, 비타민K2"},
    39: {"category": "오메가3",   "ingredient": "오메가3, 피쉬 오일"},
    40: {"category": "오메가3",   "ingredient": "오메가3, 피쉬 오일"},
    41: {"category": "마그네슘", "ingredient": "마그네슘"},
    42: {"category": "콜라겐",   "ingredient": "콜라겐 펩타이드, 히알루론산, 비타민C"},
    43: {"category": "확인필요", "ingredient": "크레아틴 일수화물"},
    44: {"category": "비타민C",   "ingredient": "비타민C"},
    45: {"category": "오메가3",   "ingredient": "오메가3, 피쉬 오일"},
    46: {"category": "비타민C",   "ingredient": "비타민C"},
    47: {"category": "오메가3",   "ingredient": "오메가3, 피쉬 오일"},
    48: {"category": "확인필요", "ingredient": "비타민D3"}
}

def main():
    if not os.path.exists(INPUT_PATH):
        print(f"Error: {INPUT_PATH}가 존재하지 않습니다.")
        return

    # 스타일 유지를 위해 복사해서 연 후 수정
    wb = load_workbook(INPUT_PATH)
    ws = wb.active
    
    # 데이터 업데이트
    for row, data in MAPPING.items():
        # 카테고리는 4번째 열 (D)
        ws.cell(row=row, column=4, value=data["category"])
        # 성분은 5번째 열 (E)
        ws.cell(row=row, column=5, value=data["ingredient"])
        
    wb.save(OUTPUT_PATH)
    print("--- 초안 엑셀 생성 완료 ---")
    print(f"저장 완료: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
