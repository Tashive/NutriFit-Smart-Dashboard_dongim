"""
추가분 엑셀 초안 작성 및 저장 스크립트

이 스크립트는 `성분_수작업입력_추가분_콜라겐유산균.xlsx` 파일을 로드하여,
카테고리 셀의 메모(Comment) 힌트를 바탕으로 카테고리를 채우고,
제품명에서 추출한 성분명(콜라겐 펩타이드, 프로바이오틱스, 비타민D 등)을 성분 셀에 기입합니다.
최종 수정된 결과는 `성분_수작업입력_추가분_초안.xlsx` 파일로 새로 저장됩니다.
"""

import os
import sys
from openpyxl import load_workbook

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
INPUT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_추가분_콜라겐유산균.xlsx")
OUTPUT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_추가분_초안.xlsx")

# 각 행(5행 ~ 34행)별 입력 정보 매핑
MAPPING = {
    # 콜라겐 상품군 (5행 ~ 19행)
    5:  {"category": "콜라겐", "ingredient": "콜라겐 펩타이드, 히알루론산, 비타민C"},
    6:  {"category": "콜라겐", "ingredient": "비변성 타입 II 콜라겐"},
    7:  {"category": "콜라겐", "ingredient": "콜라겐"},
    8:  {"category": "콜라겐", "ingredient": "콜라겐 펩타이드, 히알루론산, 비타민C"},
    9:  {"category": "콜라겐", "ingredient": "마린 콜라겐"},
    10: {"category": "콜라겐", "ingredient": "해양 콜라겐 펩타이드"},
    11: {"category": "콜라겐", "ingredient": "비변성 II형 콜라겐"},
    12: {"category": "콜라겐", "ingredient": "콜라겐"},
    13: {"category": "콜라겐", "ingredient": "해양 콜라겐 펩타이드"},
    14: {"category": "콜라겐", "ingredient": "콜라겐"},
    15: {"category": "콜라겐", "ingredient": "해양 콜라겐 펩타이드, 히알루론산"},
    16: {"category": "콜라겐", "ingredient": "저분자 콜라겐"},
    17: {"category": "콜라겐", "ingredient": "콜라겐"},
    18: {"category": "콜라겐", "ingredient": "저분자 콜라겐"},
    19: {"category": "콜라겐", "ingredient": "히알루론산, 콜라겐"},
    
    # 유산균 상품군 (20행 ~ 34행)
    20: {"category": "유산균", "ingredient": "프로바이오틱"},
    21: {"category": "유산균", "ingredient": "프로바이오틱"},
    22: {"category": "유산균", "ingredient": "크랜베리, 프로바이오틱"},
    23: {"category": "유산균", "ingredient": "프로바이오틱"},
    24: {"category": "유산균", "ingredient": "프로바이오틱"},
    25: {"category": "유산균", "ingredient": "프로바이오틱"},
    26: {"category": "유산균", "ingredient": "프로바이오틱"},
    27: {"category": "유산균", "ingredient": "프로바이오틱, 비타민D"},
    28: {"category": "유산균", "ingredient": "프로바이오틱"},
    29: {"category": "유산균", "ingredient": "프로바이오틱"},
    30: {"category": "유산균", "ingredient": "프로바이오틱"},
    31: {"category": "유산균", "ingredient": "프로바이오틱스, 비타민D"},
    32: {"category": "유산균", "ingredient": "프로바이오틱스, 비타민D"},
    33: {"category": "유산균", "ingredient": "멀티비타민, 유산균"},
    34: {"category": "유산균", "ingredient": "유산균"}
}

def main():
    if not os.path.exists(INPUT_PATH):
        print(f"Error: {INPUT_PATH}가 존재하지 않습니다.")
        return

    # 기존 스타일을 완벽하게 유지하기 위해 파일 로드
    wb = load_workbook(INPUT_PATH)
    ws = wb.active

    # 데이터 매핑 작성
    for row, data in MAPPING.items():
        ws.cell(row=row, column=4, value=data["category"]) # 카테고리 (D)
        ws.cell(row=row, column=5, value=data["ingredient"]) # 성분 (E)

    wb.save(OUTPUT_PATH)
    print("--- 추가분 초안 엑셀 생성 및 저장 완료 ---")
    print(f"저장 성공: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
