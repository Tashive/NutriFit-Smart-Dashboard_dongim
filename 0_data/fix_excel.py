"""
수작업 성분 입력용 엑셀 최종 수정 스크립트

이 스크립트는 `성분_수작업입력_초안.xlsx` 파일의 오류를 확인하고 수정합니다:
1. 헤더 위치를 4행으로 조정하고 헤더 컬럼명을 `플랫폼 / 제품명 / 가격(원) / 카테고리 / 성분`으로 설정합니다.
2. 1번째 상품 데이터를 5행으로 이동하고, 총 45개 상품 데이터가 5행~49행에 배치되도록 합니다.
3. 원본 CSV 데이터(NutriFit-Dashboard-Data-file.csv)를 바탕으로 가격 컬럼의 자릿수 오류(0이 하나 더 붙은 현상)를 올바르게 수정합니다.
4. 기존 디자인 스타일(안내 배너, 폰트, 테두리, 제브라 스트라이프, 가격 서식)을 그대로 유지 및 재적용합니다.
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
CSV_PATH = os.path.join(DATA_DIR, "NutriFit-Dashboard-Data-file.csv")
OUTPUT_PATH = os.path.join(DATA_DIR, "성분_수작업입력_초안.xlsx")

# 1차 카테고리 및 성분 매핑 (인덱스 순서대로 45개 상품)
MAPPING_DATA = [
    # Coupang (15)
    {"category": "마그네슘", "ingredient": "마그네슘"},
    {"category": "확인필요", "ingredient": "엽산, 비타민 B-12"},
    {"category": "확인필요", "ingredient": "케르세틴, 브로멜라인"},
    {"category": "확인필요", "ingredient": "확인필요"},
    {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    {"category": "확인필요", "ingredient": "비타민 B"},
    {"category": "확인필요", "ingredient": "철분"},
    {"category": "확인필요", "ingredient": "확인필요"},
    {"category": "마그네슘", "ingredient": "마그네슘 시트레이트"},
    {"category": "유산균", "ingredient": "프로바이오틱스, 유산균"},
    {"category": "확인필요", "ingredient": "베르베린"},
    {"category": "확인필요", "ingredient": "비타민 B-12"},
    {"category": "확인필요", "ingredient": "확인필요"},
    {"category": "멀티비타민", "ingredient": "비타민"},
    {"category": "확인필요", "ingredient": "L-리신"},
    
    # OliveYoung (15)
    {"category": "확인필요", "ingredient": "확인필요"},
    {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    {"category": "확인필요", "ingredient": "포타슘, 칼륨"},
    {"category": "멀티비타민", "ingredient": "비타민"},
    {"category": "비타민C", "ingredient": "비타민C"},
    {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    {"category": "확인필요", "ingredient": "확인필요"},
    {"category": "확인필요", "ingredient": "확인필요"},
    {"category": "멀티비타민", "ingredient": "비타민, 아르기닌"},
    {"category": "확인필요", "ingredient": "포타슘, 칼륨"},
    {"category": "멀티비타민", "ingredient": "비타민"},
    {"category": "멀티비타민", "ingredient": "비타민, 미네랄"},
    {"category": "비타민C", "ingredient": "비타민C"},
    {"category": "확인필요", "ingredient": "비타민C, 비타민D"},
    
    # iHerb (15)
    {"category": "유산균", "ingredient": "프로바이오틱"},
    {"category": "확인필요", "ingredient": "비타민D3, 비타민K2"},
    {"category": "마그네슘", "ingredient": "마그네슘 리시네이트 글리시네이트"},
    {"category": "마그네슘", "ingredient": "마그네슘 글리시네이트"},
    {"category": "확인필요", "ingredient": "비타민D3, 비타민K2"},
    {"category": "오메가3", "ingredient": "오메가3, 피쉬 오일"},
    {"category": "오메가3", "ingredient": "오메가3, 피쉬 오일"},
    {"category": "마그네슘", "ingredient": "마그네슘"},
    {"category": "콜라겐", "ingredient": "콜라겐 펩타이드, 히알루론산, 비타민C"},
    {"category": "확인필요", "ingredient": "크레아틴 일수화물"},
    {"category": "비타민C", "ingredient": "비타민C"},
    {"category": "오메가3", "ingredient": "오메가3, 피쉬 오일"},
    {"category": "비타민C", "ingredient": "비타민C"},
    {"category": "오메가3", "ingredient": "오메가3, 피쉬 오일"},
    {"category": "확인필요", "ingredient": "비타민D3"}
]

def clean_price_correctly(val):
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    # 소수점이 있으면 소수점 앞의 정수부만 남김 (.0 등의 자릿수 밀림 오류 해결)
    if "." in val_str:
        val_str = val_str.split(".")[0]
    cleaned = "".join([c for c in val_str if c.isdigit()])
    if cleaned:
        return int(cleaned)
    return None

def main():
    if not os.path.exists(CSV_PATH):
        print(f"Error: {CSV_PATH}가 존재하지 않습니다.")
        return

    df = pd.read_csv(CSV_PATH, low_memory=False)
    
    # 각 플랫폼별 상위 15개 상품 추출
    platforms = ["Coupang", "OliveYoung", "iHerb"]
    extracted_data = []

    for platform in platforms:
        group = df[df["platform"] == platform]
        top15 = group.head(15).copy()
        
        if platform == "Coupang":
            name_col = "product_name"
            price_col = "price"
        elif platform == "OliveYoung":
            name_col = "name"
            price_col = "price_cur"
        else: # iHerb
            name_col = "displayName"
            price_col = "discountPrice"
            
        for _, row in top15.iterrows():
            name = row.get(name_col)
            price_raw = row.get(price_col)
            price = clean_price_correctly(price_raw)
            
            extracted_data.append({
                "플랫폼": platform,
                "제품명": name,
                "가격": price
            })
            
    # 매핑 데이터와 매치하여 리스트 빌드
    final_rows = []
    for idx, item in enumerate(extracted_data):
        mapping = MAPPING_DATA[idx]
        final_rows.append({
            "플랫폼": item["플랫폼"],
            "제품명": item["제품명"],
            "가격": item["가격"],
            "카테고리": mapping["category"],
            "성분": mapping["ingredient"]
        })
        
    # openpyxl로 새로 엑셀 쓰기
    wb = Workbook()
    ws = wb.active
    ws.title = "성분 수작업 입력"
    ws.views.sheetView[0].showGridLines = True
    
    # 스타일 설정
    font_name = "맑은 고딕"
    font_normal = Font(name=font_name, size=10)
    font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    font_banner = Font(name=font_name, size=11, bold=True, color="7F3F00")
    
    fill_banner = PatternFill(start_color="FFEAD2", end_color="FFEAD2", fill_type="solid")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_medium = Side(border_style="medium", color="1F4E78")
    
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # 1행: 안내 배너
    ws.merge_cells("A1:E1")
    ws["A1"] = "★ 안내 문구: 카테고리는 [멀티비타민 / 비타민C / 오메가3 / 유산균 / 콜라겐 / 마그네슘] 중 하나로 채워주세요."
    ws["A1"].font = font_banner
    ws["A1"].fill = fill_banner
    ws["A1"].alignment = align_center
    ws["A1"].border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    ws.row_dimensions[1].height = 35
    
    # 2행: 공백
    ws.row_dimensions[2].height = 15
    
    # 3행: 공백 (4행 헤더 배치를 위한 패딩)
    ws.row_dimensions[3].height = 15
    
    # 4행: 헤더
    headers = ["플랫폼", "제품명", "가격(원)", "카테고리", "성분"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = Border(left=border_thin, right=border_thin, top=border_medium, bottom=border_medium)
    ws.row_dimensions[4].height = 28
    
    # 5행 ~ 49행: 데이터 작성 (총 45개 행)
    for i, data in enumerate(final_rows):
        r_idx = 5 + i
        use_zebra = (r_idx % 2 == 0)
        
        # 플랫폼
        c_a = ws.cell(row=r_idx, column=1, value=data["플랫폼"])
        c_a.alignment = align_center
        
        # 제품명
        c_b = ws.cell(row=r_idx, column=2, value=data["제품명"])
        c_b.alignment = align_left
        
        # 가격
        c_c = ws.cell(row=r_idx, column=3, value=data["가격"])
        c_c.alignment = align_right
        c_c.number_format = "₩#,##0"
        
        # 카테고리
        c_d = ws.cell(row=r_idx, column=4, value=data["카테고리"])
        c_d.alignment = align_center
        
        # 성분
        c_e = ws.cell(row=r_idx, column=5, value=data["성분"])
        c_e.alignment = align_center
        
        # 스타일 및 보더 일괄 적용
        for col_num in range(1, 6):
            c = ws.cell(row=r_idx, column=col_num)
            c.font = font_normal
            c.border = cell_border
            if use_zebra:
                c.fill = fill_zebra
                
        ws.row_dimensions[r_idx].height = 22
        
    # 열 너비 설정
    widths = {
        "A": 15,
        "B": 65,
        "C": 15,
        "D": 22,
        "E": 22
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save(OUTPUT_PATH)
    print("--- 최종 엑셀 수정 및 저장 완료 ---")
    print(f"저장 완료: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
