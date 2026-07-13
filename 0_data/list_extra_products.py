"""
추가분 엑셀 상품명 목록 출력 스크립트

이 스크립트는 `성분_수작업입력_추가분_콜라겐유산균.xlsx` 파일을 로드하여
각 행 번호, 플랫폼, 제품명, 그리고 카테고리 셀의 메모(Comment) 내용을 출력합니다.
분석된 제품명을 바탕으로 카테고리와 성분 값을 기입하는 스크립트를 작성하는 데 활용됩니다.
"""

import os
import sys
from openpyxl import load_workbook

# 출력 인코딩을 UTF-8로 설정
sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"c:\\Users\\tasha\\OneDrive\\바탕 화면\\ICB10_02\\project2_team3\\0_data"
EXCEL_PATH = os.path.join(DATA_DIR, "성분_수작업입력_추가분_콜라겐유산균.xlsx")

wb = load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

# 5행부터 데이터 시작 (4행은 헤더)
for row in range(5, ws.max_row + 1):
    platform = ws.cell(row=row, column=1).value
    name = ws.cell(row=row, column=2).value
    comment = ws.cell(row=row, column=4).comment
    comment_text = comment.text if comment else "None"
    print(f"Row {row:2d} | {platform} | {comment_text} | {name}")
